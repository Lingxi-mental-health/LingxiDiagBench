"""
Benchmark运行器

统一运行所有benchmark任务并生成汇总报告
"""

import json
import os
import time
from typing import Dict, List, Any, Optional
from datetime import datetime

import numpy as np

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("警告: pandas未安装，Excel输出功能将不可用")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("警告: openpyxl未安装，Excel格式化功能将受限")

try:
    from .config import (
        TRAIN_DATA_FILE, TEST_DATA_100_FILE, TEST_DATA_500_FILE,
        OUTPUT_DIR, MODEL_DIR
    )
    from .data_utils import load_and_process_data
except ImportError:
    from config import (
        TRAIN_DATA_FILE, TEST_DATA_100_FILE, TEST_DATA_500_FILE,
        OUTPUT_DIR, MODEL_DIR
    )
    from data_utils import load_and_process_data


def convert_to_serializable(obj):
    """将对象转换为可序列化的格式"""
    if isinstance(obj, (np.integer, np.int64)):
        return int(obj)
    elif isinstance(obj, (np.floating, np.float64)):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_serializable(i) for i in obj]
    return obj


class BenchmarkRunner:
    """
    Benchmark运行器
    
    统一管理和运行所有benchmark任务
    """
    
    def __init__(
        self,
        train_file: str = None,
        test_file: str = None,
        output_dir: str = None
    ):
        """
        初始化运行器
        
        Args:
            train_file: 训练数据文件路径
            test_file: 测试数据文件路径
            output_dir: 输出目录
        """
        self.train_file = train_file or TRAIN_DATA_FILE
        self.test_file = test_file or TEST_DATA_100_FILE
        self.output_dir = output_dir or OUTPUT_DIR
        
        os.makedirs(self.output_dir, exist_ok=True)
        
        self.results = {}
        self.start_time = None
    
    def run_tfidf_benchmark(
        self,
        classifier_types: List[str] = None,
        classification_types: List[str] = None
    ) -> List[Dict[str, Any]]:
        """
        运行TF-IDF benchmark
        
        Args:
            classifier_types: 分类器类型列表
            classification_types: 分类类型列表
            
        Returns:
            结果列表
        """
        try:
            from .tfidf_classifier import run_tfidf_benchmark
        except ImportError:
            from tfidf_classifier import run_tfidf_benchmark
        
        if classifier_types is None:
            classifier_types = ["logistic", "svm", "rf"]
        
        print("\n" + "="*80)
        print("运行 TF-IDF Benchmark")
        print("="*80)
        
        # 加载数据
        train_data = load_and_process_data(self.train_file)
        test_data = load_and_process_data(self.test_file)
        
        results = []
        
        for clf_type in classifier_types:
            for class_type in classification_types or ["2class", "4class", "12class"]:
                try:
                    try:
                        from .tfidf_classifier import train_and_evaluate_tfidf
                    except ImportError:
                        from tfidf_classifier import train_and_evaluate_tfidf
                    result = train_and_evaluate_tfidf(
                        train_data, test_data,
                        classification_type=class_type,
                        classifier_type=clf_type,
                        save_model=True
                    )
                    results.append(result)
                except Exception as e:
                    print(f"TF-IDF {class_type} ({clf_type}) 评估失败: {e}")
        
        self.results['tfidf'] = results
        return results
    
    def run_bert_benchmark(
        self,
        classification_types: List[str] = None
    ) -> List[Dict[str, Any]]:
        """
        运行BERT benchmark
        
        Args:
            classification_types: 分类类型列表
            
        Returns:
            结果列表
        """
        try:
            try:
                from .bert_classifier import run_bert_benchmark, TRANSFORMERS_AVAILABLE
            except ImportError:
                from bert_classifier import run_bert_benchmark, TRANSFORMERS_AVAILABLE
            if not TRANSFORMERS_AVAILABLE:
                print("跳过BERT benchmark: transformers库未安装")
                return []
        except ImportError:
            print("跳过BERT benchmark: 导入失败")
            return []
        
        print("\n" + "="*80)
        print("运行 BERT Benchmark")
        print("="*80)
        
        results = run_bert_benchmark(
            self.train_file,
            self.test_file,
            classification_types=classification_types or ["2class", "4class", "12class"]
        )
        
        self.results['bert'] = results
        return results
    
    def run_roberta_benchmark(
        self,
        classification_types: List[str] = None
    ) -> List[Dict[str, Any]]:
        """
        运行RoBERTa benchmark
        
        Args:
            classification_types: 分类类型列表
            
        Returns:
            结果列表
        """
        try:
            try:
                from .roberta_classifier import run_roberta_benchmark, TRANSFORMERS_AVAILABLE
            except ImportError:
                from roberta_classifier import run_roberta_benchmark, TRANSFORMERS_AVAILABLE
            if not TRANSFORMERS_AVAILABLE:
                print("跳过RoBERTa benchmark: transformers库未安装")
                return []
        except ImportError:
            print("跳过RoBERTa benchmark: 导入失败")
            return []
        
        print("\n" + "="*80)
        print("运行 RoBERTa Benchmark")
        print("="*80)
        
        results = run_roberta_benchmark(
            self.train_file,
            self.test_file,
            classification_types=classification_types or ["2class", "4class", "12class"]
        )
        
        self.results['roberta'] = results
        return results
    
    def run_llm_benchmark(
        self,
        model_name: str = None,
        classification_types: List[str] = None
    ) -> List[Dict[str, Any]]:
        """
        运行LLM Zero-shot benchmark
        
        Args:
            model_name: LLM模型名称
            classification_types: 分类类型列表
            
        Returns:
            结果列表
        """
        try:
            try:
                from .llm_zeroshot_classifier import run_llm_benchmark
            except ImportError:
                from llm_zeroshot_classifier import run_llm_benchmark
        except Exception as e:
            print(f"跳过LLM benchmark: {e}")
            return []
        
        print("\n" + "="*80)
        print("运行 LLM Zero-shot Benchmark")
        print("="*80)
        
        results = run_llm_benchmark(
            self.test_file,
            model_name=model_name,
            classification_types=classification_types or ["2class", "4class", "12class"]
        )
        
        self.results['llm'] = results
        return results
    
    def run_next_utterance_benchmark(
        self,
        predictions: List[str] = None,
        predictions_file: str = None,
        use_bert_score: bool = True,
        eval_interval: int = 5,
        model_name: str = None,
        max_workers: int = 32,
        api_key: str = None,
        limit: int = 5000
    ) -> Dict[str, Any]:
        """
        运行医生提问下一句预测benchmark
        
        支持三种模式：
        1. 提供 predictions 列表：直接使用提供的预测进行评估
        2. 提供 predictions_file：从文件加载预测进行评估
        3. 提供 model_name：使用 LLM 生成预测后评估
        
        Args:
            predictions: 预测结果列表（优先级最高）
            predictions_file: 预测结果文件路径（优先级次之）
            use_bert_score: 是否计算BertScore
            eval_interval: 采样间隔，每隔多少轮采样一次（默认5）
            model_name: LLM模型名称（如果提供，将使用LLM生成预测）
            max_workers: 并行工作线程数（用于LLM预测）
            api_key: API密钥（仅用于OpenRouter）
            limit: 最大样本数限制，固定 seed 随机采样（默认 5000）
            
        Returns:
            结果字典
        """
        try:
            from .next_utterance_predictor import run_next_utterance_benchmark
        except ImportError:
            from next_utterance_predictor import run_next_utterance_benchmark
        
        print("\n" + "="*80)
        print("运行 医生提问下一句预测 Benchmark")
        print("="*80)
        
        result = run_next_utterance_benchmark(
            self.test_file,
            predictions=predictions,
            predictions_file=predictions_file,
            use_bert_score=use_bert_score,
            eval_interval=eval_interval,
            model_name=model_name,
            max_workers=max_workers,
            api_key=api_key,
            limit=limit,
            output_dir=self.output_dir
        )
        
        self.results['next_utterance'] = result
        return result
    
    def run_all(
        self,
        skip_bert: bool = False,
        skip_roberta: bool = False,
        skip_llm: bool = False,
        skip_next_utterance: bool = False,
        llm_model: str = None,
        next_utterance_eval_interval: int = 5,
        next_utterance_limit: int = 5000
    ) -> Dict[str, Any]:
        """
        运行所有benchmark
        
        Args:
            skip_bert: 是否跳过BERT
            skip_roberta: 是否跳过RoBERTa
            skip_llm: 是否跳过LLM
            skip_next_utterance: 是否跳过下一句预测
            llm_model: LLM模型名称
            next_utterance_eval_interval: 医生提问预测的采样间隔
            next_utterance_limit: 医生提问预测的最大样本数限制
            
        Returns:
            所有结果
        """
        self.start_time = time.time()
        
        print("\n" + "#"*80)
        print("# 精神疾病诊断静态Benchmark")
        print(f"# 开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("#"*80)
        
        # TF-IDF
        self.run_tfidf_benchmark()
        
        # BERT
        if not skip_bert:
            self.run_bert_benchmark()
        
        # RoBERTa
        if not skip_roberta:
            self.run_roberta_benchmark()
        
        # LLM
        if not skip_llm:
            self.run_llm_benchmark(model_name=llm_model)
        
        # 下一句预测（使用与分类任务相同的 LLM 模型）
        if not skip_next_utterance:
            self.run_next_utterance_benchmark(
                eval_interval=next_utterance_eval_interval,
                model_name=llm_model,
                limit=next_utterance_limit,
                max_workers=32 # 使用32个线程进行预测
            )
        
        elapsed = time.time() - self.start_time
        
        print("\n" + "#"*80)
        print(f"# Benchmark完成！")
        print(f"# 总耗时: {elapsed:.2f}秒")
        print("#"*80)
        
        return self.results
    
    def generate_summary(self) -> Dict[str, Any]:
        """
        生成汇总报告
        
        Returns:
            汇总报告字典
        """
        summary = {
            'timestamp': datetime.now().isoformat(),
            'train_file': self.train_file,
            'test_file': self.test_file,
            'total_time': time.time() - self.start_time if self.start_time else 0,
            'classification_results': [],
            'generation_results': []
        }
        
        # 分类任务汇总
        for method in ['tfidf', 'bert', 'roberta', 'llm']:
            if method not in self.results:
                continue
            
            for result in self.results[method]:
                metrics = result.get('metrics', {})
                
                entry = {
                    'method': result.get('method', method.upper()),
                    'classification_type': result.get('classification_type', ''),
                    'classifier_type': result.get('classifier_type', ''),
                    'model_name': result.get('model_name', ''),
                }
                
                # 提取关键指标
                if result.get('classification_type') == '12class':
                    entry.update({
                        'accuracy': metrics.get('accuracy'),
                        'exact_match_accuracy': metrics.get('exact_match_accuracy'),
                        'sample_accuracy': metrics.get('sample_accuracy'),
                        'label_accuracy': metrics.get('label_accuracy'),
                        'top1_accuracy': metrics.get('top1_accuracy'),
                        'top3_accuracy': metrics.get('top3_accuracy'),
                        'hamming_accuracy': metrics.get('hamming_accuracy'),
                        'macro_precision': metrics.get('macro_precision'),
                        'macro_recall': metrics.get('macro_recall'),
                        'macro_f1': metrics.get('macro_f1'),
                        'weighted_precision': metrics.get('weighted_precision'),
                        'weighted_recall': metrics.get('weighted_recall'),
                        'weighted_f1': metrics.get('weighted_f1'),
                    })
                else:
                    entry.update({
                        'accuracy': metrics.get('accuracy'),
                        'macro_precision': metrics.get('macro_precision'),
                        'macro_recall': metrics.get('macro_recall'),
                        'macro_f1': metrics.get('macro_f1'),
                        'weighted_precision': metrics.get('weighted_precision'),
                        'weighted_recall': metrics.get('weighted_recall'),
                        'weighted_f1': metrics.get('weighted_f1'),
                    })
                
                summary['classification_results'].append(entry)
        
        # 生成任务汇总
        if 'next_utterance' in self.results:
            metrics = self.results['next_utterance'].get('metrics', {})
            summary['generation_results'].append({
                'task': 'next_utterance_prediction',
                'bleu': metrics.get('bleu'),
                'bleu-1': metrics.get('bleu-1'),
                'bleu-4': metrics.get('bleu-4'),
                'rouge_l_f1': metrics.get('rouge_l_f1'),
                'bert_score_f1': metrics.get('bert_score_f1'),
            })
        
        return summary
    
    def save_results(self, filename: str = None, model_name: str = None) -> str:
        """
        保存结果到JSON文件
        
        Args:
            filename: 输出文件名
            
        Returns:
            输出文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if model_name:
                filename = f"benchmark_results_{model_name}_{timestamp}.json"
            else:
                filename = f"benchmark_results_{timestamp}.json"
        
        output_path = os.path.join(self.output_dir, filename)
        
        # 转换为可序列化格式
        results_serializable = convert_to_serializable(self.results)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results_serializable, f, ensure_ascii=False, indent=2)
        
        print(f"\n结果已保存到: {output_path}")
        return output_path
    
    def export_to_excel(self, filename: str = None) -> str:
        """
        导出结果到Excel文件
        
        Args:
            filename: 输出文件名
            
        Returns:
            输出文件路径
        """
        if not PANDAS_AVAILABLE:
            print("错误: pandas未安装，无法导出Excel")
            return None
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"benchmark_results_{timestamp}.xlsx"
        
        output_path = os.path.join(self.output_dir, filename)
        
        summary = self.generate_summary()
        
        # 创建分类结果DataFrame
        clf_df = pd.DataFrame(summary['classification_results'])
        
        # 创建生成任务结果DataFrame
        gen_df = pd.DataFrame(summary['generation_results'])
        
        # 写入Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 分类结果
            clf_df.to_excel(writer, sheet_name='分类任务结果', index=False)
            
            # 生成任务结果
            if not gen_df.empty:
                gen_df.to_excel(writer, sheet_name='生成任务结果', index=False)
            
            # 详细结果
            for method in ['tfidf', 'bert', 'roberta', 'llm']:
                if method not in self.results:
                    continue
                
                for i, result in enumerate(self.results[method]):
                    class_type = result.get('classification_type', '')
                    clf_type = result.get('classifier_type', '')
                    
                    # 类别详细指标
                    class_metrics = result.get('metrics', {}).get('class_metrics', {})
                    if class_metrics:
                        class_df = pd.DataFrame([
                            {
                                'class': cls,
                                'precision': m.get('precision', 0),
                                'recall': m.get('recall', 0),
                                'f1': m.get('f1', 0),
                                'support': m.get('support', 0)
                            }
                            for cls, m in class_metrics.items()
                        ])
                        
                        sheet_name = f"{method}_{class_type}"
                        if clf_type:
                            sheet_name += f"_{clf_type}"
                        sheet_name = sheet_name[:31]  # Excel sheet名称限制
                        
                        class_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\nExcel结果已保存到: {output_path}")
        
        # 格式化Excel
        if OPENPYXL_AVAILABLE:
            self._format_excel(output_path)
        
        return output_path
    
    def _format_excel(self, filepath: str):
        """格式化Excel文件"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = load_workbook(filepath)
        
        # 定义样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet in wb.worksheets:
            # 格式化标题行
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 格式化数据行
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    # 格式化百分比
                    if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                        cell.number_format = '0.00%'
            
            # 自动调整列宽
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        print(f"Excel格式化完成")
    
    def export_summary_excel(
        self, 
        filename: str = "benchmark_summary.xlsx",
        append: bool = True
    ) -> str:
        """
        导出汇总结果到Excel（按图示格式，支持追加）
        
        格式：
        Method/Model | 2class(Acc,F1-macro,F1-weighted) | 4class(...) | 12class(Acc,Top1,Top3,F1-macro,F1-weighted)
        
        Args:
            filename: 输出文件名
            append: 是否追加到现有文件
            
        Returns:
            输出文件路径
        """
        if not PANDAS_AVAILABLE:
            print("错误: pandas未安装，无法导出Excel")
            return None
        
        output_path = os.path.join(self.output_dir, filename)
        
        # 定义列名（按图示格式）
        columns = [
            'Method/Model',
            '2class_Acc', '2class_F1-macro', '2class_F1-weighted',
            '4class_Acc', '4class_F1-macro', '4class_F1-weighted',
            '12class_Acc', '12class_Top1_Acc', '12class_Top3_Acc', '12class_F1-macro', '12class_F1-weighted'
        ]
        
        # 收集当前运行的结果
        new_rows = self._collect_summary_rows()
        
        if not new_rows:
            print("没有结果需要导出")
            return None
        
        # 如果文件存在且需要追加
        if append and os.path.exists(output_path):
            try:
                existing_df = pd.read_excel(output_path)
                # 确保列名一致
                for col in columns:
                    if col not in existing_df.columns:
                        existing_df[col] = None
                existing_df = existing_df[columns]
            except Exception as e:
                print(f"读取现有文件失败: {e}，将创建新文件")
                existing_df = pd.DataFrame(columns=columns)
        else:
            existing_df = pd.DataFrame(columns=columns)
        
        # 创建新行的DataFrame
        new_df = pd.DataFrame(new_rows, columns=columns)
        
        # 合并
        result_df = pd.concat([existing_df, new_df], ignore_index=True)
        
        # 保存
        result_df.to_excel(output_path, index=False, engine='openpyxl')
        
        # 格式化
        if OPENPYXL_AVAILABLE:
            self._format_summary_excel(output_path)
        
        print(f"\n汇总结果已{'追加' if append else '保存'}到: {output_path}")
        return output_path
    
    def _collect_summary_rows(self) -> List[List]:
        """收集汇总行数据"""
        rows = []
        
        # 按方法分组收集结果
        method_results = {}
        
        for method in ['tfidf', 'bert', 'roberta', 'llm']:
            if method not in self.results:
                continue
            
            for result in self.results[method]:
                clf_type = result.get('classifier_type', '')
                model_name = result.get('model_name', '')
                class_type = result.get('classification_type', '')
                metrics = result.get('metrics', {})
                
                # 生成方法名称
                if method == 'tfidf':
                    method_name = f"TF-IDF + {clf_type}"
                elif method == 'bert':
                    method_name = "Bert"
                elif method == 'roberta':
                    method_name = "RoBERTa"
                else:
                    # LLM: 提取模型名称
                    if model_name:
                        # 从 "qwen3-32b:9041" 提取 "Qwen3-32B"
                        name_part = model_name.split(':')[0]
                        method_name = name_part.replace('-', '-').title().replace('Qwen', 'Qwen')
                    else:
                        method_name = "LLM"
                
                # 初始化该方法的结果字典
                if method_name not in method_results:
                    method_results[method_name] = {
                        '2class': {}, '4class': {}, '12class': {}
                    }
                
                # 提取指标
                if class_type == '2class':
                    method_results[method_name]['2class'] = {
                        'acc': metrics.get('accuracy', 0),
                        'f1_macro': metrics.get('macro_f1', 0),
                        'f1_weighted': metrics.get('weighted_f1', 0)
                    }
                elif class_type == '4class':
                    method_results[method_name]['4class'] = {
                        'acc': metrics.get('accuracy', 0),
                        'f1_macro': metrics.get('macro_f1', 0),
                        'f1_weighted': metrics.get('weighted_f1', 0)
                    }
                elif class_type == '12class':
                    method_results[method_name]['12class'] = {
                        'acc': metrics.get('accuracy', metrics.get('exact_match_accuracy', 0)),
                        'top1_acc': metrics.get('top1_accuracy', 0),
                        'top3_acc': metrics.get('top3_accuracy', 0),
                        'f1_macro': metrics.get('macro_f1', 0),
                        'f1_weighted': metrics.get('weighted_f1', 0)
                    }
        
        # 转换为行数据
        for method_name, class_results in method_results.items():
            row = [method_name]
            
            # 2class
            r2 = class_results.get('2class', {})
            row.extend([
                r2.get('acc'),
                r2.get('f1_macro'),
                r2.get('f1_weighted')
            ])
            
            # 4class
            r4 = class_results.get('4class', {})
            row.extend([
                r4.get('acc'),
                r4.get('f1_macro'),
                r4.get('f1_weighted')
            ])
            
            # 12class
            r12 = class_results.get('12class', {})
            row.extend([
                r12.get('acc'),
                r12.get('top1_acc'),
                r12.get('top3_acc'),
                r12.get('f1_macro'),
                r12.get('f1_weighted')
            ])
            
            rows.append(row)
        
        return rows
    
    def _format_summary_excel(self, filepath: str):
        """格式化汇总Excel文件"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 定义样式
        header_font = Font(bold=True, color='000000')
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 格式化标题行
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 格式化数据行
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # 格式化百分比（0-1之间的浮点数）
                if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                    cell.number_format = '0.00%'
        
        # 自动调整列宽
        for i, column in enumerate(ws.columns):
            max_length = 0
            column_letter = get_column_letter(i + 1)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def export_next_utterance_excel(
        self,
        filename: str = "next_utterance_summary.xlsx",
        model_name: str = None,
        append: bool = True
    ) -> str:
        """
        导出医生提问下一句预测结果到Excel（支持追加）
        
        格式：
        Method/Model | BLEU | Rouge-L | BertScore | Total_Samples
        
        Args:
            filename: 输出文件名
            model_name: 模型名称（用于标识行）
            append: 是否追加到现有文件
            
        Returns:
            输出文件路径
        """
        if not PANDAS_AVAILABLE:
            print("错误: pandas未安装，无法导出Excel")
            return None
        
        if 'next_utterance' not in self.results:
            print("没有next_utterance结果需要导出")
            return None
        
        output_path = os.path.join(self.output_dir, filename)
        
        # 定义列名（简化版：只保留核心指标）
        columns = [
            'Method/Model',
            'BLEU',
            'Rouge-L',
            'BertScore',
            'Total_Samples'
        ]
        
        # 收集当前运行的结果
        result = self.results['next_utterance']
        metrics = result.get('metrics', {})
        
        # 确定模型名称
        if model_name is None:
            model_name = "Default"
        
        new_row = [
            model_name,
            metrics.get('bleu'),
            metrics.get('rouge_l_f1'),      # 使用 Rouge-L F1 作为 Rouge-L 指标
            metrics.get('bert_score_f1'),   # 使用 BertScore F1 作为 BertScore 指标
            metrics.get('total_samples'),
        ]
        
        # 如果文件存在且需要追加
        if append and os.path.exists(output_path):
            try:
                existing_df = pd.read_excel(output_path)
                # 确保列名一致
                for col in columns:
                    if col not in existing_df.columns:
                        existing_df[col] = None
                existing_df = existing_df[columns]
            except Exception as e:
                print(f"读取现有文件失败: {e}，将创建新文件")
                existing_df = pd.DataFrame(columns=columns)
        else:
            existing_df = pd.DataFrame(columns=columns)
        
        # 创建新行的DataFrame
        new_df = pd.DataFrame([new_row], columns=columns)
        
        # 合并
        result_df = pd.concat([existing_df, new_df], ignore_index=True)
        
        # 保存
        result_df.to_excel(output_path, index=False, engine='openpyxl')
        
        # 格式化
        if OPENPYXL_AVAILABLE:
            self._format_next_utterance_excel(output_path)
        
        print(f"\nNext Utterance结果已{'追加' if append else '保存'}到: {output_path}")
        return output_path
    
    def _format_next_utterance_excel(self, filepath: str):
        """格式化Next Utterance Excel文件"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 定义样式
        header_font = Font(bold=True, color='000000')
        header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 格式化标题行
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 格式化数据行
        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # 格式化百分比（0-1之间的浮点数，排除Total_Samples列）
                col_name = ws.cell(row=1, column=i+1).value
                if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                    if col_name not in ['Total_Samples']:
                        cell.number_format = '0.00%'
        
        # 自动调整列宽
        for i, column in enumerate(ws.columns):
            max_length = 0
            column_letter = get_column_letter(i + 1)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='精神疾病诊断静态Benchmark')
    parser.add_argument('--train-file', type=str, default=TRAIN_DATA_FILE,
                        help='训练数据文件路径')
    parser.add_argument('--test-file', type=str, default=TEST_DATA_100_FILE,
                        help='测试数据文件路径')
    parser.add_argument('--output-dir', type=str, default=OUTPUT_DIR,
                        help='输出目录')
    parser.add_argument('--skip-bert', action='store_true',
                        help='跳过BERT benchmark')
    parser.add_argument('--skip-llm', action='store_true',
                        help='跳过LLM benchmark')
    parser.add_argument('--skip-next-utterance', action='store_true',
                        help='跳过下一句预测benchmark')
    parser.add_argument('--llm-model', type=str, default='qwen3-32b:9041',
                        help='LLM模型名称')
    parser.add_argument('--tfidf-only', action='store_true',
                        help='只运行TF-IDF benchmark')
    
    args = parser.parse_args()
    
    # 创建运行器
    runner = BenchmarkRunner(
        train_file=args.train_file,
        test_file=args.test_file,
        output_dir=args.output_dir
    )
    
    if args.tfidf_only:
        runner.run_tfidf_benchmark()
    else:
        runner.run_all(
            skip_bert=args.skip_bert,
            skip_llm=args.skip_llm,
            skip_next_utterance=args.skip_next_utterance,
            llm_model=args.llm_model
        )
    
    # 保存结果
    runner.save_results()
    runner.export_to_excel()


if __name__ == "__main__":
    main()

