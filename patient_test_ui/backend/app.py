#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Patient Agent 测试 UI - 后端 API 服务
提供与 Patient LLM 交互的 RESTful API 接口

主要功能：
1. 加载患者数据
2. 初始化Patient Agent 
3. 处理多轮对话
4. 生成诊断结果
"""

import os
import sys
import json
import time
from typing import Dict, List, Any, Optional
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import traceback
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# 添加项目根目录到 Python 路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# 导入项目模块
try:
    from config import CONFIG, PROJECT_ROOT
    from session_manager import PatientSession
    from utils import safe_get_text, safe_truncate_text
    from auth import UserStore
    from agents import AutoDiagnosisVerifier
except ImportError as e:
    print(f"导入模块失败: {e}")
    print(f"当前路径: {os.getcwd()}")
    print(f"项目根目录: {project_root}")
    sys.exit(1)

project_root = Path(PROJECT_ROOT)

app = Flask(__name__, static_folder='../frontend')
CORS(app)  # 允许跨域请求
USER_STORE = UserStore(CONFIG["user_db_path"])

# 全局变量
patient_data: List[Dict] = []
active_sessions: Dict[str, PatientSession] = {}  # session_id -> session
_diagnosis_agent = None
_diagnosis_agent_disabled = False


def _get_diagnosis_agent() -> AutoDiagnosisVerifier:
    global _diagnosis_agent, _diagnosis_agent_disabled
    if _diagnosis_agent_disabled:
        raise RuntimeError("自动诊断模型不可用")
    if _diagnosis_agent is None:
        try:
            _diagnosis_agent = AutoDiagnosisVerifier(
                CONFIG["models"]["verifier"],
                CONFIG["openrouter"],
            )
        except Exception as exc:
            _diagnosis_agent_disabled = True
            print(f"初始化自动诊断代理失败: {exc}")
            traceback.print_exc()
            raise RuntimeError("自动诊断模型初始化失败") from exc
    return _diagnosis_agent


def _normalize_role(raw: Any) -> str:
    if raw is None:
        return "system"
    text = str(raw).strip().lower()
    mapping = {
        "doctor": "doctor",
        "医生": "doctor",
        "医师": "doctor",
        "patient": "patient",
        "患者": "patient",
        "病人": "patient",
        "患者本人": "patient",
        "system": "system",
        "系统": "system",
        "unknown": "unknown",
        "未知发言人": "unknown",
        "family": "family",
        "家属": "family",
        "患者家属": "family",
        "others": "unknown",
    }
    return mapping.get(text, "unknown")


def _normalize_content(raw: Any) -> str:
    if raw is None:
        return ""
    return safe_truncate_text(str(raw).strip(), 4000)


def _extract_conversation_from_sheet(sheet) -> List[Dict[str, str]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    role_index: Optional[int] = None
    text_index: Optional[int] = None

    for idx, name in enumerate(header):
        if name in {"role", "speaker", "角色", "身份"} and role_index is None:
            role_index = idx
        if name in {"content", "message", "text", "对话", "内容", "话语"} and text_index is None:
            text_index = idx

    if text_index is None:
        text_index = 1 if len(header) > 1 else 0

    conversation: List[Dict[str, str]] = []
    for row in rows[1:]:
        if row is None:
            continue
        role_val = None if role_index is None or role_index >= len(row) else row[role_index]
        text_val = None if text_index is None or text_index >= len(row) else row[text_index]
        normalized_text = _normalize_content(text_val)
        if not normalized_text:
            continue
        conversation.append(
            {
                "role": _normalize_role(role_val),
                "content": normalized_text,
            }
        )

    return conversation


@app.route('/api/auth/users', methods=['GET'])
def list_users():
    """列出所有已注册的用户名"""
    try:
        return jsonify({
            "success": True,
            "data": USER_STORE.list_users()
        })
    except Exception as e:
        print(f"获取用户列表失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/auth/register', methods=['POST'])
def register_user():
    """注册新用户"""
    try:
        data = request.get_json() or {}
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""

        USER_STORE.register(username, password)

        return jsonify({
            "success": True,
            "message": "注册成功",
            "data": {
                "username": username
            }
        })
    except ValueError as ve:
        return jsonify({
            "success": False,
            "error": str(ve)
        }), 400
    except Exception as e:
        print(f"注册用户失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/auth/login', methods=['POST'])
def login_user():
    """验证用户登录"""
    try:
        data = request.get_json() or {}
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""

        if not USER_STORE.authenticate(username, password):
            return jsonify({
                "success": False,
                "error": "用户名或密码错误"
            }), 401

        return jsonify({
            "success": True,
            "data": {
                "username": username
            }
        })
    except Exception as e:
        print(f"用户登录失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


def load_patient_data() -> bool:
    """加载患者数据"""
    global patient_data
    
    try:
        with open(CONFIG["patient_data_path"], 'r', encoding='utf-8') as f:
            all_data = json.load(f)
        
        # 限制患者数量以提高性能
        patient_data = all_data[:CONFIG["max_patients"]]
        
        # 数据格式转换，统一字段名
        for i, patient in enumerate(patient_data):
            # 处理SMHC_EverDiag-16K_validation_data_100samples.json的字段格式
            if "patient_id" in patient:
                # 新格式数据，直接使用现有字段，并转换为整数
                try:
                    patient["患者"] = int(patient.get("patient_id", i + 1))
                except (ValueError, TypeError):
                    patient["患者"] = i + 1
                # 保持现有字段不变，Patient Agent会直接使用
                if "Patient info" not in patient:
                    # 安全处理，过滤掉None值和敏感信息
                    filtered_patient_info = {
                        k: v for k, v in patient.items() 
                        if v is not None and k not in ["Diagnosis", "DiagnosisCode", "OverallDiagnosis"]
                    }
                    patient["Patient info"] = json.dumps(filtered_patient_info, ensure_ascii=False)
                if "cleaned_text" not in patient:
                    patient["cleaned_text"] = safe_get_text(patient, "PresentIllnessHistory")
            else:
                # 旧格式数据兼容处理
                if "患者" not in patient:
                    patient["患者"] = i + 1
                
                # 创建Patient Agent需要的格式
                patient["patient_id"] = patient.get("患者", i + 1)
                patient["Age"] = patient.get("年龄", 0)
                patient["Gender"] = patient.get("性别", "未知")
                patient["Department"] = patient.get("科室", "精神科")
                patient["ChiefComplaint"] = patient.get("主诉", "")
                patient["PresentIllnessHistory"] = patient.get("现病史", "")
                patient["Patient info"] = json.dumps(patient, ensure_ascii=False)
                patient["cleaned_text"] = patient.get("现病史", "")
        
        print(f"成功加载 {len(patient_data)} 个患者数据")
        return True
        
    except Exception as e:
        print(f"加载患者数据失败: {e}")
        traceback.print_exc()
        return False


# API 路由

@app.route('/data/<path:filename>')
def serve_data_files(filename):
    """提供data目录下的数据文件"""
    import os
    data_dir = os.path.join(os.path.dirname(__file__), '../data')
    return send_from_directory(data_dir, filename)

@app.route('/')
def index():
    """服务主页"""
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:filename>')
def static_files(filename):
    """静态文件服务"""
    return send_from_directory(app.static_folder, filename)

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        "status": "healthy",
        "timestamp": time.time(),
        "config": {
            "model_name": CONFIG["openrouter_model"] if CONFIG["use_openrouter"] else CONFIG["model_name"],
            "model_port": CONFIG["model_port"],
            "patients_loaded": len(patient_data),
            "use_openrouter": CONFIG["use_openrouter"],
            "openrouter_model": CONFIG["openrouter_model"] if CONFIG["use_openrouter"] else None,
            "doctor_model": CONFIG["models"]["doctor"],
            "verifier_model": CONFIG["models"]["verifier"],
            "agent_versions": CONFIG["agent_versions"]
        }
    })

@app.route('/api/patients', methods=['GET'])
def get_patients():
    """获取患者列表"""
    try:
        username = request.args.get('username')
        completed_patients = set()
        annotated_patients = set()
        
        if username:
            # 扫描用户标注记录
            user_dir = project_root / "patient_test_ui" / "user_annotations" / username
            if user_dir.exists():
                for annotation_file in user_dir.glob("annotation_*.json"):
                    try:
                        with open(annotation_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            patient_id = str(data.get("patient_id", ""))
                            if patient_id:
                                annotated_patients.add(patient_id)
                    except Exception:
                        continue
            
            # 扫描评测记录，找出该用户已完成的患者
            evaluation_dir = project_root / "patient_test_ui" / "evaluations"
            if evaluation_dir.exists():
                for eval_file in evaluation_dir.glob("evaluation_*.json"):
                    try:
                        with open(eval_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            if data.get("evaluator") == username:
                                completed_patients.add(str(data.get("patient_id")))
                    except Exception:
                        continue

        # 返回患者基本信息（隐藏真实诊断）
        patients_list = []
        for patient in patient_data:
            p_id = str(patient.get("患者", patient.get("patient_id", 0)))
            patients_list.append({
                "patient_id": p_id,
                "age": patient.get("Age", patient.get("年龄", 0)),
                "gender": safe_get_text(patient, "Gender", "性别", "未知"),
                "department": safe_get_text(patient, "Department", "科室", "精神科"),
                "chief_complaint": safe_truncate_text(
                    safe_get_text(patient, "ChiefComplaint", "主诉"), 100
                ),
                "completed": p_id in completed_patients,
                "annotated": p_id in annotated_patients
                # 不再返回真实诊断信息
                # "diagnosis": patient.get("诊断结果", ""),
                # "icd_code": patient.get("ICD编码", "")
            })
        
        # 按patient_id的整数值从小到大排序
        def get_patient_id_int(patient_item):
            try:
                return int(patient_item["patient_id"])
            except (ValueError, TypeError):
                return 0  # 如果转换失败，返回0作为默认值
        
        patients_list.sort(key=get_patient_id_int)
        
        return jsonify({
            "success": True,
            "data": patients_list,
            "total": len(patients_list)
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/doctors', methods=['GET'])
def get_doctors():
    """获取医生列表"""
    try:
        import json
        import os
        
        # 读取医生角色配置文件
        doctor_persona_path = os.path.join(os.path.dirname(__file__), '../../prompts/doctor/doctor_persona.json')
        
        with open(doctor_persona_path, 'r', encoding='utf-8') as f:
            doctors = json.load(f)
        
        # 为每个医生添加ID
        for i, doctor in enumerate(doctors):
            doctor['doctor_id'] = i + 1
        
        return jsonify({
            "success": True,
            "data": doctors,
            "count": len(doctors)
        })
        
    except Exception as e:
        print(f"获取医生列表失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/patients/<patient_id>', methods=['GET'])
def get_patient_detail(patient_id):
    """获取患者详细信息（隐藏真实诊断）"""
    try:
        # 支持字符串和整数类型的patient_id
        def match_patient_id(patient):
            p_id = patient.get("患者", patient.get("patient_id"))
            # 将两个值都转换为字符串进行比较，以处理类型不匹配的问题
            return str(p_id) == str(patient_id)
        
        patient = next((p for p in patient_data if match_patient_id(p)), None)
        if not patient:
            return jsonify({
                "success": False,
                "error": f"患者 {patient_id} 不存在"
            }), 404
        
        return jsonify({
            "success": True,
            "data": {
                "patient_id": str(patient.get("患者", patient.get("patient_id", 0))),
                "age": patient.get("Age", patient.get("年龄", 0)),
                "gender": safe_get_text(patient, "Gender", "性别", "未知"),
                "department": safe_get_text(patient, "Department", "科室", "精神科"),
                "chief_complaint": safe_get_text(patient, "ChiefComplaint", "主诉"),
                "present_illness": safe_get_text(patient, "PresentIllnessHistory", "现病史"),
                # 不再返回真实诊断信息，保持隐藏
                # "diagnosis": patient.get("诊断结果", ""),
                # "icd_code": patient.get("ICD编码", ""),
                "physical_disease_history": safe_get_text(patient, "ImportantRelevantPhysicalIllnessHistory", "重要或相关躯体疾病史"),
                "family_history": safe_get_text(patient, "FamilyHistory", "精神疾病家族史"),
                "personal_history": safe_get_text(patient, "PersonalHistory", "个人史")
            }
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/sessions', methods=['POST'])
def create_session():
    """创建新的对话会话"""
    try:
        data = request.get_json()
        patient_id = data.get("patient_id")
        doctor_id = data.get("doctor_id")
        user_name = (data.get("user_name") or data.get("username") or "").strip()
        patient_version = data.get("patient_version", CONFIG["agent_versions"]["patient"]["default"])
        doctor_version = data.get("doctor_version", CONFIG["agent_versions"]["doctor"]["default"])

        if not user_name:
            return jsonify({
                "success": False,
                "error": "缺少 user_name 参数"
            }), 400

        # 如果提供了医生ID，创建医生问诊会话
        if doctor_id is not None:
            return create_doctor_session(doctor_id, user_name, doctor_version)

        if not patient_id:
            return jsonify({
                "success": False,
                "error": "缺少 patient_id 参数"
            }), 400

        if user_name not in USER_STORE.list_users():
            return jsonify({
                "success": False,
                "error": "用户不存在，请先注册"
            }), 400
        
        # 验证patient_version
        available_patient_versions = CONFIG["agent_versions"]["patient"]["available"]
        if patient_version not in available_patient_versions:
            return jsonify({
                "success": False,
                "error": f"不支持的Patient版本: {patient_version}。可用版本: {', '.join(available_patient_versions)}"
            }), 400
        
        # 查找患者数据
        def match_patient_id(patient):
            p_id = patient.get("患者", patient.get("patient_id"))
            # 将两个值都转换为字符串进行比较，以处理类型不匹配的问题
            return str(p_id) == str(patient_id)
        
        patient = next((p for p in patient_data if match_patient_id(p)), None)
        if not patient:
            return jsonify({
                "success": False,
                "error": f"患者 {patient_id} 不存在"
            }), 404
        
        # 创建会话ID
        session_id = f"session_{patient_id}_{int(time.time())}"
        
        # 创建会话实例（传入版本参数）
        session = PatientSession(patient, session_id, user_name=user_name, patient_version=patient_version)
        
        # 初始化Patient Agent
        if not session.initialize_patient():
            return jsonify({
                "success": False,
                "error": "初始化Patient Agent失败"
            }), 500
        
        # 存储会话
        active_sessions[session_id] = session
        session.record_event(
            "session_created",
            {
                "patient_info": session.get_session_info()["patient_info"],
                "patient_version": patient_version,
            },
        )
        
        # 不再自动生成患者主诉，等待医生先开始对话
        
        return jsonify({
            "success": True,
            "data": {
                "session_id": session_id,
                "patient_info": session.get_session_info()["patient_info"],
                "user_name": user_name,
                "patient_version": patient_version,
                # 移除 chief_complaint 字段
            }
        })
        
    except Exception as e:
        print(f"创建会话失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


def create_doctor_session(doctor_id, user_name, doctor_version="base"):
    """创建医生问诊会话"""
    try:
        import json
        import os
        
        # 验证doctor_version
        available_doctor_versions = CONFIG["agent_versions"]["doctor"]["available"]
        if doctor_version not in available_doctor_versions:
            return jsonify({
                "success": False,
                "error": f"不支持的Doctor版本: {doctor_version}。可用版本: {', '.join(available_doctor_versions)}"
            }), 400
        
        # 读取医生角色配置文件
        doctor_persona_path = os.path.join(os.path.dirname(__file__), '../../prompts/doctor/doctor_persona.json')
        
        with open(doctor_persona_path, 'r', encoding='utf-8') as f:
            doctors = json.load(f)
        
        # 查找指定的医生
        if doctor_id < 1 or doctor_id > len(doctors):
            return jsonify({
                "success": False,
                "error": f"未找到医生ID: {doctor_id}"
            }), 404
        
        doctor = doctors[doctor_id - 1]  # 数组索引从0开始
        doctor['doctor_id'] = doctor_id
        
        # 创建会话ID
        session_id = f"doctor_session_{doctor_id}_{int(time.time())}"
        
        # 创建医生会话对象（传入版本参数）
        from session_manager import DoctorSession
        session = DoctorSession(
            session_id=session_id,
            doctor_data=doctor,
            user_name=user_name,
            doctor_version=doctor_version
        )
        
        active_sessions[session_id] = session
        
        return jsonify({
            "success": True,
            "data": {
                "session_id": session_id,
                "session_type": "doctor",
                "doctor_version": doctor_version,
                "doctor_info": {
                    "doctor_id": doctor_id,
                    "name": doctor.get("name"),
                    "age": doctor.get("age"),
                    "gender": doctor.get("gender"),
                    "special": doctor.get("special"),
                    "speed": doctor.get("speed"),
                    "commu": doctor.get("commu"),
                    "empathy": doctor.get("empathy"),
                    "explain": doctor.get("explain")
                }
            }
        })
        
    except Exception as e:
        print(f"创建医生会话失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "创建医生会话失败"
        }), 500


@app.route('/api/sessions/<session_id>/chat', methods=['POST'])
def chat_with_patient(session_id):
    """与患者对话"""
    try:
        data = request.get_json()
        doctor_message = data.get("message", "").strip()
        
        if not doctor_message:
            return jsonify({
                "success": False,
                "error": "医生消息不能为空"
            }), 400
        
        # 检查会话是否存在
        if session_id not in active_sessions:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404
        
        session = active_sessions[session_id]
        
        # 获取患者回复
        result = session.patient_response(doctor_message)
        
        if result["error"]:
            return jsonify({
                "success": False,
                "error": result["error"]
            }), 500
        
        return jsonify({
            "success": True,
            "data": {
                "patient_response": result["response"],
                "classification": result["classification"],
                "session_info": session.get_session_info(),
                "suggested_questions": result.get("suggested_questions", [])
            }
        })
        
    except Exception as e:
        print(f"对话处理失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/sessions/<session_id>/doctor-question', methods=['POST'])
def generate_doctor_question(session_id):
    """生成医生问题，供用户扮演患者回答"""
    try:
        session = active_sessions.get(session_id)
        if not session:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404

        result = session.generate_doctor_question()

        return jsonify({
            "success": True,
            "data": {
                "question": result["question"],
                "is_diagnosis": result.get("is_diagnosis", False),
                "session_info": session.get_session_info(),
            }
        })

    except ValueError as exc:
        return jsonify({
            "success": False,
            "error": str(exc)
        }), 400
    except Exception as exc:
        print(f"生成医生问题失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "生成医生问题失败"
        }), 500


@app.route('/api/sessions/<session_id>/patient-reply', methods=['POST'])
def record_patient_reply(session_id):
    """记录扮演患者的手动回答"""
    try:
        session = active_sessions.get(session_id)
        if not session:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404

        data = request.get_json() or {}
        reply = (data.get("message") or "").strip()
        if not reply:
            return jsonify({
                "success": False,
                "error": "患者回答不能为空"
            }), 400

        # 检查会话类型，调用相应的方法
        if hasattr(session, 'record_patient_reply'):
            # 医生会话
            session.record_patient_reply(reply)
        else:
            # 传统患者会话
            session.record_patient_reply_manual(reply)

        return jsonify({
            "success": True,
            "data": {
                "session_info": session.get_session_info(),
            }
        })

    except ValueError as exc:
        return jsonify({
            "success": False,
            "error": str(exc)
        }), 400
    except Exception as exc:
        print(f"记录患者回答失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "记录患者回答失败"
        }), 500


@app.route('/api/sessions/<session_id>', methods=['GET'])
def get_session_info(session_id):
    """获取会话信息"""
    try:
        if session_id not in active_sessions:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404
        
        session = active_sessions[session_id]
        
        return jsonify({
            "success": True,
            "data": session.get_session_info()
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/sessions/<session_id>/recommend-question', methods=['POST'])
def recommend_question(session_id):
    """推荐下一轮可提问的问题或患者回答"""
    try:
        session = active_sessions.get(session_id)
        if not session:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404

        request_payload = request.get_json(silent=True) or {}
        mode = request_payload.get("mode", "doctor_question")  # 默认为医生问题模式
        
        try:
            count = int(request_payload.get("count", 1))
        except (TypeError, ValueError):
            count = 1

        # 根据模式选择不同的推荐功能
        if mode == "patient_answer":
            # 患者回答推荐模式 (EverPsychosis)
            if hasattr(session, 'recommend_patient_answers'):
                # DoctorSession - 推荐患者回答
                has_doctor_question = any(log.get("role") == "doctor" for log in session.conversation_log)
                if not has_doctor_question:
                    return jsonify({
                        "success": False,
                        "error": "请等待医生提问后再获取推荐回答"
                    }), 400
                
                answers = session.recommend_patient_answers(limit=count)
                return jsonify({
                    "success": True,
                    "data": {
                        "questions": answers  # 保持前端兼容性，仍使用questions字段
                    }
                })
            else:
                return jsonify({
                    "success": False,
                    "error": "当前会话类型不支持患者回答推荐"
                }), 400
        else:
            # 医生问题推荐模式 (EverPsychiatrist)
            if hasattr(session, 'recommend_questions'):
                # PatientSession - 推荐医生问题
                has_patient_reply = any(log.get("role") == "patient" for log in session.conversation_log)
                if not has_patient_reply:
                    return jsonify({
                        "success": False,
                        "error": "请先完成至少一轮问诊后再获取推荐问题"
                    }), 400
                
                questions = session.recommend_questions(limit=count)
                return jsonify({
                    "success": True,
                    "data": {
                        "questions": questions
                    }
                })
            else:
                return jsonify({
                    "success": False,
                    "error": "当前会话类型不支持医生问题推荐"
                }), 400

    except Exception as e:
        print(f"推荐问题失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/sessions/<session_id>/auto-diagnosis', methods=['POST'])
def auto_diagnosis(session_id):
    """自动生成诊断结果"""
    try:
        session = active_sessions.get(session_id)
        if not session:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404

        try:
            result = session.run_auto_diagnosis()
        except ValueError as ve:
            return jsonify({
                "success": False,
                "error": str(ve)
            }), 400
        except Exception as e:
            print(f"自动诊断失败: {e}")
            traceback.print_exc()
            return jsonify({
                "success": False,
                "error": f"自动诊断失败: {e}"
            }), 500

        return jsonify({
            "success": True,
            "data": result
        })
    except Exception as e:
        print(f"自动诊断接口异常: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


def _extract_patient_info_from_sheet(sheet) -> List[Dict[str, Any]]:
    """从Excel表格中提取病人信息"""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    
    # 查找相关列的索引
    patient_id_index = None
    age_index = None
    gender_index = None
    conversation_index = None
    
    for idx, name in enumerate(header):
        name_lower = name.lower()
        if name_lower in {"visitnumber", "visit_number", "patient_id", "患者id", "患者编号", "id"} and patient_id_index is None:
            patient_id_index = idx
        elif name_lower in {"age", "年龄"} and age_index is None:
            age_index = idx
        elif name_lower in {"gender", "性别"} and gender_index is None:
            gender_index = idx
        elif name_lower in {"cleaned_text", "conversation", "对话", "对话内容", "conversation_data"} and conversation_index is None:
            conversation_index = idx

    patients = []
    for row_idx, row in enumerate(rows[1:], 1):
        if row is None:
            continue
            
        # 提取病人基本信息
        patient_id = None
        if patient_id_index is not None and patient_id_index < len(row):
            patient_id = row[patient_id_index]
        if patient_id is None:
            patient_id = row_idx  # 使用行号作为默认ID
            
        age = None
        if age_index is not None and age_index < len(row):
            try:
                age = int(row[age_index]) if row[age_index] is not None else None
            except (ValueError, TypeError):
                age = None
                
        gender = None
        if gender_index is not None and gender_index < len(row):
            gender = str(row[gender_index]).strip() if row[gender_index] is not None else None
            
        # 提取对话内容
        conversation_data = None
        if conversation_index is not None and conversation_index < len(row):
            conversation_data = str(row[conversation_index]).strip() if row[conversation_index] is not None else None

        patients.append({
            "patient_id": patient_id,
            "age": age,
            "gender": gender,
            "conversation_data": conversation_data,
            "row_number": row_idx
        })

    return patients


def _extract_data_from_json(json_data, include_conversation: bool = True) -> tuple[List[Dict[str, str]], List[Dict[str, Any]]]:
    """从JSON数据中提取对话记录和病人信息
    
    Args:
        json_data: JSON数据列表
        include_conversation: 是否包含对话内容（优化：大量数据时设为False）
    """
    conversation = []
    patients = []
    
    # 确保json_data是列表格式
    if not isinstance(json_data, list):
        raise ValueError("JSON文件应该包含一个患者数据的数组")
    
    for idx, patient_record in enumerate(json_data):
        if not isinstance(patient_record, dict):
            continue
            
        # 提取患者基本信息
        patient_id = patient_record.get("patient_id", idx + 1)
        age = patient_record.get("Age")
        gender = patient_record.get("Gender")
        diagnosis = patient_record.get("Diagnosis", "")
        diagnosis_code = patient_record.get("DiagnosisCode", "")
        
        # 提取对话内容（如果需要）
        cleaned_text = patient_record.get("cleaned_text", "")
        
        # 解析对话内容为结构化格式（只在需要时处理）
        if include_conversation and cleaned_text:
            patient_conversation = _parse_conversation_content(cleaned_text)
            conversation.extend(patient_conversation)
        
        # 构建患者信息（不包含完整对话内容，只保留标识）
        patient_info = {
            "patient_id": patient_id,
            "age": age,
            "gender": gender,
            "conversation_data": cleaned_text if include_conversation else "",  # 优化：不返回对话文本
            "has_conversation": bool(cleaned_text),  # 标记是否有对话数据
            "diagnosis": diagnosis,
            "diagnosis_code": diagnosis_code,
            "chief_complaint": patient_record.get("ChiefComplaint", ""),
            "present_illness_history": patient_record.get("PresentIllnessHistory", ""),
            "personal_history": patient_record.get("PersonalHistory", ""),
            "family_history": patient_record.get("FamilyHistory", ""),
            "department": patient_record.get("Department", ""),
            "row_number": idx + 1
        }
        
        patients.append(patient_info)
    
    return conversation, patients


@app.route('/api/diagnosis/import-excel', methods=['POST'])
def import_diagnosis_excel():
    """Parse uploaded Excel file and extract conversation records and patient info."""
    try:
        if 'file' not in request.files:
            return jsonify({
                "success": False,
                "error": "缺少上传文件"
            }), 400

        upload = request.files['file']
        if not upload.filename:
            return jsonify({
                "success": False,
                "error": "请选择需要上传的 Excel 文件"
            }), 400

        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active
        
        # 提取对话记录（原有功能）
        conversation = _extract_conversation_from_sheet(sheet)
        
        # 提取病人信息（新功能）
        patients = _extract_patient_info_from_sheet(sheet)

        return jsonify({
            "success": True,
            "data": {
                "conversation": conversation,
                "records": len(conversation),
                "sheet_name": sheet.title,
                "patients": patients,
                "patient_count": len(patients)
            }
        })

    except InvalidFileException:
        return jsonify({
            "success": False,
            "error": "无法读取该文件，请上传有效的 Excel 文档"
        }), 400
    except Exception as exc:
        print(f"解析 Excel 失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "导入对话记录失败"
        }), 500


@app.route('/api/diagnosis/import-json', methods=['POST'])
def import_diagnosis_json():
    """Parse uploaded JSON file and extract conversation records and patient info."""
    try:
        if 'file' not in request.files:
            return jsonify({
                "success": False,
                "error": "缺少上传文件"
            }), 400

        upload = request.files['file']
        if not upload.filename:
            return jsonify({
                "success": False,
                "error": "请选择需要上传的 JSON 文件"
            }), 400

        # 读取JSON文件内容
        json_content = upload.read().decode('utf-8')
        json_data = json.loads(json_content)
        
        # 提取对话记录和病人信息
        conversation, patients = _extract_data_from_json(json_data)

        return jsonify({
            "success": True,
            "data": {
                "conversation": conversation,
                "records": len(conversation),
                "file_name": upload.filename,
                "patients": patients,
                "patient_count": len(patients)
            }
        })

    except json.JSONDecodeError as exc:
        print(f"JSON 解析失败: {exc}")
        return jsonify({
            "success": False,
            "error": "无法解析 JSON 文件，请检查文件格式"
        }), 400
    except Exception as exc:
        print(f"解析 JSON 失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "导入对话记录失败"
        }), 500


# 全局缓存：缓存默认数据文件内容，避免重复读取
_default_data_cache = {
    "data": None,
    "file_path": None,
    "load_time": None
}

@app.route('/api/diagnosis/load-default-data', methods=['GET'])
def load_default_diagnosis_data():
    """Load default diagnosis data from configured patient data path.
    优化：只返回病人列表，不返回对话内容，大幅提升加载速度
    """
    try:
        # 使用CONFIG中配置的数据文件路径
        default_data_path = Path(CONFIG["patient_data_path"])
        
        if not default_data_path.exists():
            return jsonify({
                "success": False,
                "error": f"默认数据文件不存在: {default_data_path}"
            }), 404

        # 读取JSON文件内容（使用缓存）
        if (_default_data_cache["file_path"] == str(default_data_path) and 
            _default_data_cache["data"] is not None):
            print(f"使用缓存的数据文件: {default_data_path}")
            json_data = _default_data_cache["data"]
        else:
            print(f"正在读取数据文件: {default_data_path}")
            with open(default_data_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            # 缓存数据
            _default_data_cache["data"] = json_data
            _default_data_cache["file_path"] = str(default_data_path)
            _default_data_cache["load_time"] = time.time()
            print(f"已缓存 {len(json_data)} 条数据")
        
        # 应用max_patients限制到原始数据
        max_patients = CONFIG.get("max_patients", len(json_data))
        limited_json_data = json_data[:max_patients]
        
        # 提取病人信息（不包含对话内容，大幅提升速度）
        _, patients = _extract_data_from_json(limited_json_data, include_conversation=False)

        return jsonify({
            "success": True,
            "data": {
                "conversation": [],  # 不再返回所有对话，改为按需加载
                "records": 0,
                "file_name": os.path.basename(default_data_path),
                "patients": patients,
                "patient_count": len(patients),
                "total_patients": len(json_data),
                "using_cache": _default_data_cache["file_path"] is not None
            }
        })

    except json.JSONDecodeError as exc:
        print(f"默认JSON数据解析失败: {exc}")
        return jsonify({
            "success": False,
            "error": "无法解析默认JSON数据文件，请检查文件格式"
        }), 500
    except Exception as exc:
        print(f"加载默认数据失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "加载默认诊断数据失败"
        }), 500


@app.route('/api/diagnosis/patient-conversation', methods=['GET'])
def get_patient_conversation():
    """获取单个病人的对话内容（按需加载，提升性能）"""
    try:
        patient_id = request.args.get('patient_id')
        if not patient_id:
            return jsonify({
                "success": False,
                "error": "缺少patient_id参数"
            }), 400
        
        # 使用缓存的数据
        if _default_data_cache["data"] is None:
            # 如果缓存为空，重新加载
            default_data_path = Path(CONFIG["patient_data_path"])
            if not default_data_path.exists():
                return jsonify({
                    "success": False,
                    "error": "数据文件不存在"
                }), 404
            
            with open(default_data_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            _default_data_cache["data"] = json_data
            _default_data_cache["file_path"] = str(default_data_path)
        else:
            json_data = _default_data_cache["data"]
        
        # 查找对应的病人数据
        patient_record = None
        for record in json_data:
            if str(record.get("patient_id", "")) == str(patient_id):
                patient_record = record
                break
        
        if not patient_record:
            return jsonify({
                "success": False,
                "error": f"未找到patient_id为{patient_id}的病人"
            }), 404
        
        # 提取对话内容
        cleaned_text = patient_record.get("cleaned_text", "")
        
        if not cleaned_text:
            return jsonify({
                "success": True,
                "data": {
                    "conversation": [],
                    "patient_id": patient_id
                }
            })
        
        # 解析对话内容
        conversation = _parse_conversation_content(cleaned_text)
        
        return jsonify({
            "success": True,
            "data": {
                "conversation": conversation,
                "patient_id": patient_id,
                "conversation_text": cleaned_text
            }
        })
        
    except Exception as exc:
        print(f"获取病人对话失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "获取病人对话失败"
        }), 500


@app.route('/api/diagnosis/generate', methods=['POST'])
def generate_diagnosis_from_records():
    """Generate diagnosis reasoning and conclusion from conversation records."""
    try:
        data = request.get_json() or {}
        conversation = data.get("conversation") or []
        patient_id = data.get("patient_id")
        username = data.get("username")
        
        print(f"=== 诊断生成API请求 ===")
        print(f"收到对话记录数量: {len(conversation)}")
        print(f"患者ID: {patient_id}, 用户: {username}")
        
        # 首先尝试从缓存加载
        if patient_id:
            # 1. 检查用户标注缓存
            if username:
                user_annotation = _load_user_annotation(username, patient_id)
                if user_annotation:
                    print(f"从用户标注缓存加载诊断数据")
                    return jsonify({
                        "success": True,
                        "data": user_annotation,
                        "from_cache": "user_annotation"
                    })
            
            # 2. 检查预加载缓存
            cached_diagnosis = _load_cached_diagnosis(patient_id)
            if cached_diagnosis:
                print(f"从预加载缓存加载诊断数据")
                return jsonify({
                    "success": True,
                    "data": cached_diagnosis,
                    "from_cache": "preload"
                })
        
        # 没有缓存，进行实时生成
        if not isinstance(conversation, list) or not conversation:
            return jsonify({
                "success": False,
                "error": "请先导入对话记录"
            }), 400

        formatted_conversation = [
            {
                "role": _normalize_role(item.get("role")),
                "content": _normalize_content(item.get("content")),
            }
            for item in conversation
            if item and _normalize_content(item.get("content"))
        ]

        print(f"格式化后的对话: {json.dumps(formatted_conversation, ensure_ascii=False, indent=2)}")

        if not formatted_conversation:
            return jsonify({
                "success": False,
                "error": "对话记录为空"
            }), 400

        agent = _get_diagnosis_agent()
        result = agent.generate_diagnosis(formatted_conversation)
        
        print(f"=== AI模型返回结果 ===")
        print(f"完整结果: {json.dumps(result, ensure_ascii=False, indent=2)}")
        print(f"原始返回: {result.get('raw', '')}")
        print(f"提取的思考过程: {result.get('thought', '')}")
        print(f"提取的reasoning: {result.get('reasoning', '')}")
        print(f"提取的ICD代码: {result.get('icd_codes', [])}")
        print(f"使用的模型: {result.get('model', '')}")
        print(f"========================")

        return jsonify({
            "success": True,
            "data": result,
            "from_cache": False
        })

    except RuntimeError as exc:
        return jsonify({
            "success": False,
            "error": str(exc)
        }), 503
    except Exception as exc:
        print(f"自动诊断生成失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "自动生成诊断失败"
        }), 500


def _load_cached_diagnosis(patient_id: str) -> Optional[Dict[str, Any]]:
    """从预加载缓存加载诊断数据"""
    # 从环境变量读取缓存文件路径，如果没有设置则使用默认路径
    cache_file_path = os.getenv(
        "DIAGNOSIS_CACHE_FILE",
        str(project_root / "patient_test_ui" / "data" / "diagnosis_cache.json")
    )
    cache_file = Path(cache_file_path)
    
    if not cache_file.exists():
        print(f"诊断缓存文件不存在: {cache_file}")
        return None
    
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
        
        patient_id_str = str(patient_id)
        if patient_id_str in cache_data:
            cached = cache_data[patient_id_str]
            diagnosis_result = cached.get("diagnosis_result", {})
            
            print(f"从预加载缓存加载诊断数据: {cache_file}")
            
            # 转换为API返回格式
            return {
                "thought": diagnosis_result.get("thought", ""),
                "reasoning": diagnosis_result.get("reasoning", ""),
                "icd_codes": diagnosis_result.get("icd_codes", []),
                "icd_box": diagnosis_result.get("icd_box", ""),
                "model": diagnosis_result.get("model", "cached"),
                "raw": diagnosis_result.get("thought", ""),
            }
    except Exception as e:
        print(f"加载预加载缓存失败 ({cache_file}): {e}")
        traceback.print_exc()
    
    return None


def _load_user_annotation(username: str, patient_id: str) -> Optional[Dict[str, Any]]:
    """加载用户的标注数据"""
    user_dir = project_root / "patient_test_ui" / "user_annotations" / username
    
    if not user_dir.exists():
        return None
    
    try:
        # 查找该患者的最新标注
        patient_id_str = str(patient_id)
        annotations = []
        
        for file in user_dir.glob(f"annotation_{patient_id_str}_*.json"):
            try:
                with open(file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if data.get("patient_id") == patient_id_str:
                        annotations.append(data)
            except Exception:
                continue
        
        if not annotations:
            return None
        
        # 返回最新的标注
        latest = max(annotations, key=lambda x: x.get("saved_at", 0))
        
        # 转换为API返回格式
        diagnosis = latest.get("diagnosis", {})
        return {
            "thought": diagnosis.get("reason", ""),
            "reasoning": diagnosis.get("reason", ""),
            "icd_codes": [],  # 从conclusion中提取
            "icd_box": diagnosis.get("conclusion", ""),
            "model": "user_annotation",
            "raw": diagnosis.get("reason", ""),
            "doctor_notes": latest.get("doctor_notes", ""),
            "corrected_diagnosis": latest.get("corrected_diagnosis"),
        }
    except Exception as e:
        print(f"加载用户标注失败: {e}")
        traceback.print_exc()
    
    return None


def _parse_conversation_content(conversation_text: str) -> List[Dict[str, str]]:
    """解析对话内容，按发言人分类，支持带数字的角色前缀（如"医生1："、"家属2："）"""
    if not conversation_text:
        return []
    
    import re
    
    # 分割对话内容
    lines = conversation_text.split('\n')
    parsed_conversation = []
    
    # 定义角色识别的正则表达式模式
    # 匹配 "角色" + 0个或多个数字 + "中文冒号或英文冒号"
    doctor_pattern = re.compile(r'^(医生\d*[：:])\s*(.*)', re.IGNORECASE)
    patient_pattern = re.compile(r'^(患者\d*|病人\d*|患者本人\d*)[：:]\s*(.*)', re.IGNORECASE)
    family_pattern = re.compile(r'^(家属\d*|患者家属\d*)[：:]\s*(.*)', re.IGNORECASE)
    unknown_pattern = re.compile(r'^(未知发言人\d*)[：:]\s*(.*)', re.IGNORECASE)
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # 检查是否以发言人标识开头
        role = "unknown"
        content = line
        
        # 尝试匹配各种角色模式
        match = doctor_pattern.match(line)
        if match:
            role = "doctor"
            content = match.group(2).strip()
        else:
            match = patient_pattern.match(line)
            if match:
                role = "patient"
                content = match.group(2).strip()
            else:
                match = family_pattern.match(line)
                if match:
                    role = "family"
                    content = match.group(2).strip()
                else:
                    match = unknown_pattern.match(line)
                    if match:
                        role = "unknown"
                        content = match.group(2).strip()
        
        if content:
            parsed_conversation.append({
                "role": role,
                "content": content
            })
    
    return parsed_conversation


@app.route('/api/diagnosis/parse-conversation', methods=['POST'])
def parse_conversation():
    """解析对话内容，按发言人分类显示"""
    try:
        data = request.get_json() or {}
        conversation_text = data.get("conversation_text", "").strip()
        
        if not conversation_text:
            return jsonify({
                "success": False,
                "error": "对话内容不能为空"
            }), 400
        
        parsed_conversation = _parse_conversation_content(conversation_text)
        
        return jsonify({
            "success": True,
            "data": {
                "conversation": parsed_conversation,
                "total_messages": len(parsed_conversation)
            }
        })
        
    except Exception as exc:
        print(f"解析对话内容失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "解析对话内容失败"
        }), 500


@app.route('/api/sessions/<session_id>/conversation-log', methods=['GET'])
def get_conversation_log(session_id):
    """获取会话的对话记录文件"""
    try:
        session = active_sessions.get(session_id)
        if not session:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404
        
        # 检查是否是医生会话
        if hasattr(session, 'log_filepath'):
            import os
            if os.path.exists(session.log_filepath):
                import json
                with open(session.log_filepath, 'r', encoding='utf-8') as f:
                    log_data = json.load(f)
                
                return jsonify({
                    "success": True,
                    "data": {
                        "log_file": session.log_filename,
                        "log_data": log_data,
                        "total_events": len(log_data.get("events", [])),
                        "session_info": session.get_session_info()
                    }
                })
            else:
                return jsonify({
                    "success": False,
                    "error": "对话记录文件不存在"
                }), 404
        else:
            return jsonify({
                "success": False,
                "error": "此会话类型不支持对话记录"
            }), 400
            
    except Exception as e:
        print(f"获取对话记录失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "获取对话记录失败"
        }), 500


@app.route('/api/diagnosis/save', methods=['POST'])
def save_diagnosis_result():
    """保存诊断结果（JSONL格式，同一用户的标注保存在一个文件中）"""
    try:
        data = request.get_json() or {}
        
        # 获取保存数据
        patient_id = data.get("patient_id")
        diagnosis_reason = data.get("diagnosis_reason", "").strip()
        diagnosis_conclusion = data.get("diagnosis_conclusion", "").strip()
        conversation = data.get("conversation", [])
        username = data.get("username", "anonymous")
        
        # 获取医生备注和诊断修正
        doctor_notes = data.get("doctor_notes", "").strip()
        corrected_diagnosis = data.get("corrected_diagnosis")
        
        # 获取AI生成的原始版本
        ai_generated_reason = data.get("ai_generated_reason", "").strip()
        ai_generated_conclusion = data.get("ai_generated_conclusion", "").strip()
        ai_generated_at = data.get("ai_generated_at", "")
        
        if not diagnosis_reason or not diagnosis_conclusion:
            return jsonify({
                "success": False,
                "error": "诊断原因和结论不能为空"
            }), 400
        
        # 创建保存记录
        save_record = {
            "patient_id": patient_id,
            "username": username,
            "conversation": conversation,
            "saved_at": time.time(),
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "versions": {
                "ai_generated": {
                    "diagnosis_reason": ai_generated_reason,
                    "diagnosis_conclusion": ai_generated_conclusion,
                    "generated_at": ai_generated_at
                },
                "doctor_edited": {
                    "diagnosis_reason": diagnosis_reason,
                    "diagnosis_conclusion": diagnosis_conclusion,
                    "edited_at": time.strftime("%Y-%m-%d %H:%M:%S")
                }
            },
            "doctor_notes": doctor_notes,
            "corrected_diagnosis": corrected_diagnosis
        }
        
        # 保存到用户专属的JSONL文件
        user_annotations_dir = project_root / "patient_test_ui" / "user_annotations"
        user_annotations_dir.mkdir(exist_ok=True)
        
        # JSONL文件名：用户名.jsonl
        jsonl_filename = f"{username}.jsonl"
        jsonl_file = user_annotations_dir / jsonl_filename
        
        # 追加保存到JSONL文件
        with open(jsonl_file, 'a', encoding='utf-8') as f:
            f.write(json.dumps(save_record, ensure_ascii=False) + '\n')
        
        return jsonify({
            "success": True,
            "data": {
                "saved_to": str(jsonl_file),
                "filename": jsonl_filename,
                "patient_id": patient_id,
                "timestamp": save_record["timestamp"]
            }
        })
        
    except Exception as exc:
        print(f"保存诊断结果失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "保存诊断结果失败"
        }), 500


@app.route('/api/diagnosis/load-annotation', methods=['GET'])
def load_diagnosis_annotation():
    """加载用户对特定病人的最新标注记录"""
    try:
        username = request.args.get('username', 'anonymous')
        patient_id = request.args.get('patient_id')
        
        if not patient_id:
            return jsonify({
                "success": False,
                "error": "缺少patient_id参数"
            }), 400
        
        # 读取用户的JSONL文件
        user_annotations_dir = project_root / "patient_test_ui" / "user_annotations"
        jsonl_file = user_annotations_dir / f"{username}.jsonl"
        
        if not jsonl_file.exists():
            return jsonify({
                "success": True,
                "data": None,
                "message": "未找到该用户的标注记录"
            })
        
        # 读取JSONL文件，找到该病人的最新记录（最后一条）
        latest_annotation = None
        with open(jsonl_file, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    record = json.loads(line)
                    if str(record.get('patient_id')) == str(patient_id):
                        # 如果找到匹配的病人ID，更新为最新记录
                        latest_annotation = record
        
        if latest_annotation:
            return jsonify({
                "success": True,
                "data": latest_annotation
            })
        else:
            return jsonify({
                "success": True,
                "data": None,
                "message": "未找到该病人的标注记录"
            })
            
    except Exception as exc:
        print(f"加载标注记录失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "加载标注记录失败"
        }), 500


@app.route('/api/diagnosis/annotated-patients', methods=['GET'])
def get_annotated_patients():
    """获取用户已标注的病人ID列表"""
    try:
        username = request.args.get('username', 'anonymous')
        
        # 读取用户的JSONL文件
        user_annotations_dir = project_root / "patient_test_ui" / "user_annotations"
        jsonl_file = user_annotations_dir / f"{username}.jsonl"
        
        if not jsonl_file.exists():
            return jsonify({
                "success": True,
                "data": {
                    "patient_ids": []
                }
            })
        
        # 读取所有病人ID（使用集合去重）
        patient_ids = set()
        with open(jsonl_file, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    try:
                        record = json.loads(line)
                        patient_id = record.get('patient_id')
                        if patient_id:
                            patient_ids.add(str(patient_id))
                    except json.JSONDecodeError:
                        continue
        
        return jsonify({
            "success": True,
            "data": {
                "patient_ids": list(patient_ids)
            }
        })
            
    except Exception as exc:
        print(f"获取已标注病人列表失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "获取已标注病人列表失败"
        }), 500


@app.route('/api/diagnosis/save-details', methods=['POST'])
def save_diagnosis_details():
    """保存诊断详细信息（包括结构化详情和医生修改）"""
    try:
        data = request.get_json() or {}
        
        # 获取保存数据
        patient_id = data.get("patient_id")
        diagnosis_id = data.get("diagnosis_id")
        conversation = data.get("conversation", [])
        username = data.get("username", "anonymous")
        
        # 基础诊断信息
        diagnosis_reason = data.get("diagnosis_reason", "").strip()
        diagnosis_conclusion = data.get("diagnosis_conclusion", "").strip()
        
        # 医生输入的额外信息
        doctor_notes = data.get("doctor_notes", "").strip()
        
        # 诊断修正信息（可选）
        corrected_diagnosis = data.get("corrected_diagnosis")  # {icd_code, diagnosis_name}
        
        if not diagnosis_reason or not diagnosis_conclusion:
            return jsonify({
                "success": False,
                "error": "诊断原因和结论不能为空"
            }), 400
        
        # 创建保存记录
        save_record = {
            "patient_id": patient_id,
            "diagnosis_id": diagnosis_id or f"{patient_id}_{int(time.time())}",
            "username": username,
            "conversation": conversation,
            "saved_at": time.time(),
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "diagnosis": {
                "reason": diagnosis_reason,
                "conclusion": diagnosis_conclusion,
            },
            "doctor_notes": doctor_notes,
            "corrected_diagnosis": corrected_diagnosis,
        }
        
        # 保存到用户专属目录
        user_diagnosis_dir = project_root / "patient_test_ui" / "user_annotations" / username
        user_diagnosis_dir.mkdir(parents=True, exist_ok=True)
        
        diagnosis_id = save_record["diagnosis_id"]
        filename = f"annotation_{patient_id}_{int(time.time())}.json"
        save_file = user_diagnosis_dir / filename
        
        with open(save_file, 'w', encoding='utf-8') as f:
            json.dump(save_record, f, indent=2, ensure_ascii=False)
        
        return jsonify({
            "success": True,
            "data": {
                "saved_to": str(save_file),
                "filename": filename,
                "diagnosis_id": diagnosis_id,
                "timestamp": save_record["timestamp"]
            }
        })
        
    except Exception as exc:
        print(f"保存诊断详情失败: {exc}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "保存诊断详情失败"
        }), 500


@app.route('/api/sessions/<session_id>/events', methods=['POST'])
def record_session_event(session_id):
    """记录用户交互事件（例如选择推荐问题）"""
    try:
        session = active_sessions.get(session_id)
        if not session:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404

        payload = request.get_json() or {}
        event_type = payload.get("event_type")
        event_data = payload.get("payload", {})

        if not event_type:
            return jsonify({
                "success": False,
                "error": "缺少 event_type 参数"
            }), 400

        if event_type == "suggestion_selected":
            question = (event_data or {}).get("question", "").strip()
            if not question:
                return jsonify({
                    "success": False,
                    "error": "缺少 question 信息"
                }), 400
            session.record_suggestion_selection(question)
        else:
            # 对于未明确定义的事件，拒绝记录
            return jsonify({
                "success": False,
                "error": f"不支持的事件类型: {event_type}"
            }), 400

        return jsonify({"success": True})

    except Exception as e:
        print(f"记录交互事件失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/sessions/<session_id>/diagnosis', methods=['POST'])
def create_diagnosis(session_id):
    """生成诊断结果"""
    try:
        data = request.get_json()
        diagnosis_text = data.get("diagnosis", "").strip()
        icd_code = data.get("icd_code", "").strip()
        reasoning = data.get("reasoning", "").strip()
        
        if session_id not in active_sessions:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404
        
        session = active_sessions[session_id]
        
        # 创建诊断记录(完整版,包含ground truth,仅用于保存到文件)
        diagnosis_record_full = {
            "session_id": session_id,
            "patient_id": session.patient_template.get("患者"),
            "diagnosis": diagnosis_text,
            "icd_code": icd_code,
            "reasoning": reasoning,
            "conversation_log": session.conversation_log,
            "duration": time.time() - session.start_time,
            "timestamp": time.time(),
            "actual_diagnosis": safe_get_text(session.patient_template, "Diagnosis", "诊断结果"),
            "actual_icd_code": safe_get_text(session.patient_template, "DiagnosisCode", "ICD编码")
        }
        
        # 创建诊断记录(返回给前端的版本,不包含ground truth,避免泄露)
        diagnosis_record_response = {
            "session_id": session_id,
            "patient_id": session.patient_template.get("患者"),
            "diagnosis": diagnosis_text,
            "icd_code": icd_code,
            "reasoning": reasoning,
            "conversation_log": session.conversation_log,
            "duration": time.time() - session.start_time,
            "timestamp": time.time()
            # 注意: 不包含 actual_diagnosis 和 actual_icd_code
        }
        
        # 保存诊断记录到文件(使用完整版,包含ground truth)
        diagnosis_dir = project_root / "patient_test_ui" / "diagnoses"
        diagnosis_dir.mkdir(exist_ok=True)
        
        # 使用时间戳确保每次诊断都有唯一的文件名
        diagnosis_file = diagnosis_dir / f"diagnosis_{session_id}_{int(time.time())}.json"
        with open(diagnosis_file, 'w', encoding='utf-8') as f:
            json.dump(diagnosis_record_full, f, indent=2, ensure_ascii=False)

        session.record_event(
            "manual_diagnosis_saved",
            {
                "diagnosis": diagnosis_text,
                "icd_code": icd_code,
                "reasoning": reasoning,
                "saved_to": str(diagnosis_file),
            },
        )
        
        # 返回给前端的版本不包含ground truth
        return jsonify({
            "success": True,
            "data": {
                "diagnosis_record": diagnosis_record_response,
                "saved_to": str(diagnosis_file)
            }
        })
        
    except Exception as e:
        print(f"生成诊断失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/sessions/<session_id>/evaluation', methods=['POST'])
def create_evaluation(session_id):
    """保存患者评测结果"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "未接收到评测数据"
            }), 400
        
        # 验证会话存在
        if session_id not in active_sessions:
            return jsonify({
                "success": False,
                "error": "会话不存在或已过期"
            }), 404
        
        session = active_sessions[session_id]
        
        # 构建评测记录
        evaluation_record = {
            "session_id": session_id,
            "patient_id": session.patient_template.get("患者"),
            "timestamp": time.time(),
            "evaluator": getattr(session, 'user_name', 'anonymous'),
            "dimensions": {
                "clinical_realism": {
                    "score": data.get("clinical_realism", {}).get("score"),
                    "comment": data.get("clinical_realism", {}).get("comment", "")
                },
                "interaction": {
                    "score": data.get("interaction", {}).get("score"),
                    "comment": data.get("interaction", {}).get("comment", "")
                },
                "consistency": {
                    "score": data.get("consistency", {}).get("score"),
                    "comment": data.get("consistency", {}).get("comment", "")
                },
                "safety": {
                    "score": data.get("safety", {}).get("score"),
                    "comment": data.get("safety", {}).get("comment", "")
                },
                "overall": {
                    "score": data.get("overall", {}).get("score"),
                    "comment": data.get("overall", {}).get("comment", "")
                }
            },
            "conversation_log": session.conversation_log,
            "conversation_rounds": len([log for log in session.conversation_log if log.get("role") == "patient"]),
            "duration": time.time() - session.start_time
        }
        
        # 计算平均分
        scores = [
            evaluation_record["dimensions"]["clinical_realism"]["score"],
            evaluation_record["dimensions"]["interaction"]["score"],
            evaluation_record["dimensions"]["consistency"]["score"],
            evaluation_record["dimensions"]["safety"]["score"],
            evaluation_record["dimensions"]["overall"]["score"]
        ]
        evaluation_record["average_score"] = sum(scores) / len(scores) if scores else 0
        
        # 保存评测记录到文件
        evaluation_dir = project_root / "patient_test_ui" / "evaluations"
        evaluation_dir.mkdir(exist_ok=True)
        
        # 使用时间戳确保每次评测都有唯一的文件名
        evaluation_file = evaluation_dir / f"evaluation_{session_id}_{int(time.time())}.json"
        with open(evaluation_file, 'w', encoding='utf-8') as f:
            json.dump(evaluation_record, f, indent=2, ensure_ascii=False)
        
        # 记录事件
        session.record_event(
            "patient_evaluation_submitted",
            {
                "average_score": evaluation_record["average_score"],
                "saved_to": str(evaluation_file),
                "dimensions": {
                    k: v["score"] for k, v in evaluation_record["dimensions"].items()
                }
            },
        )
        
        return jsonify({
            "success": True,
            "data": {
                "evaluation_record": evaluation_record,
                "saved_to": str(evaluation_file),
                "average_score": evaluation_record["average_score"]
            }
        })
        
    except Exception as e:
        print(f"保存评测失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/sessions/<session_id>', methods=['DELETE'])
def delete_session(session_id):
    """删除会话"""
    try:
        if session_id in active_sessions:
            session = active_sessions[session_id]
            session.record_event(
                "session_closed",
                {
                    "duration": time.time() - session.start_time,
                    "conversation_rounds": len(
                        [log for log in session.conversation_log if log.get("role") == "patient"]
                    ),
                },
            )
            del active_sessions[session_id]
        
        return jsonify({
            "success": True,
            "message": f"会话 {session_id} 已删除"
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/diagnoses', methods=['GET'])
def get_diagnoses():
    """获取所有疾病诊断和对应的ICD代码"""
    try:
        # 从ICD-10精神疾病诊断文件中读取完整的诊断列表
        icd_file_path = project_root / "patient_test_ui" / "data" / "icd10_mental_disorders.json"
        
        if not icd_file_path.exists():
            # 如果ICD文件不存在，回退到从患者数据中提取
            print(f"ICD文件不存在: {icd_file_path}，回退到患者数据提取")
            return get_diagnoses_from_patient_data()
        
        with open(icd_file_path, 'r', encoding='utf-8') as f:
            diagnoses_list = json.load(f)
        
        # 按ICD代码排序
        diagnoses_list.sort(key=lambda x: x['icd_code'])
        
        return jsonify({
            "success": True,
            "data": diagnoses_list,
            "total": len(diagnoses_list),
            "source": "icd10_file"
        })
        
    except Exception as e:
        print(f"从ICD文件获取诊断列表失败: {e}")
        traceback.print_exc()
        # 出错时回退到患者数据提取
        return get_diagnoses_from_patient_data()


def get_diagnoses_from_patient_data():
    """从患者数据中提取诊断列表（回退方案）"""
    try:
        # 从患者数据中提取所有疾病和ICD代码
        diagnoses_dict = {}
        
        for patient in patient_data:
            diagnosis = safe_get_text(patient, "Diagnosis", "诊断结果")
            icd_code = safe_get_text(patient, "DiagnosisCode", "ICD编码")
            
            # 清理诊断名称（去除末尾的逗号和空格）
            if diagnosis:
                diagnosis = diagnosis.rstrip(',').strip()
                # 处理多个诊断的情况
                diagnosis_list = [d.strip() for d in diagnosis.split(',') if d.strip()]
                icd_list = [i.strip() for i in icd_code.split(',') if i.strip()] if icd_code else []
                
                # 确保诊断和ICD代码一一对应
                for i, diag in enumerate(diagnosis_list):
                    if diag and diag not in diagnoses_dict:
                        # 使用对应位置的ICD代码，如果没有则使用第一个或空字符串
                        if i < len(icd_list):
                            diagnoses_dict[diag] = icd_list[i]
                        elif icd_list:
                            diagnoses_dict[diag] = icd_list[0]
                        else:
                            diagnoses_dict[diag] = ""
        
        # 转换为列表格式，并按诊断名称排序
        diagnoses_list = [
            {"diagnosis": diagnosis, "icd_code": icd_code}
            for diagnosis, icd_code in sorted(diagnoses_dict.items())
        ]
        
        return jsonify({
            "success": True,
            "data": diagnoses_list,
            "total": len(diagnoses_list),
            "source": "patient_data"
        })
        
    except Exception as e:
        print(f"从患者数据获取诊断列表失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/diagnoses/grouped', methods=['GET'])
def get_diagnoses_grouped():
    """获取按大类分组的ICD-10诊断编码（三级目录结构）"""
    try:
        # 优先使用层级结构的JSON文件
        hierarchical_file = project_root / "patient_test_ui" / "data" / "icd10_mental_disorders_hierarchical.json"
        
        if hierarchical_file.exists():
            with open(hierarchical_file, 'r', encoding='utf-8') as f:
                hierarchical_data = json.load(f)
            
            return jsonify({
                "success": True,
                "data": hierarchical_data["categories"],
                "total_categories": len(hierarchical_data["categories"]),
                "source": "hierarchical_file"
            })
        
        # 回退方案:从扁平文件生成层级结构
        icd_file_path = project_root / "patient_test_ui" / "data" / "icd10_mental_disorders.json"
        
        if not icd_file_path.exists():
            return jsonify({
                "success": False,
                "error": "ICD-10数据文件不存在"
            }), 404
        
        with open(icd_file_path, 'r', encoding='utf-8') as f:
            diagnoses_list = json.load(f)
        
        # 按大类分组（提取ICD代码的前缀，如F32, F33等）
        grouped_data = {}
        category_names = {
            "F00-F09": "器质性精神障碍",
            "F10-F19": "精神活性物质所致精神和行为障碍",
            "F20-F29": "精神分裂症、分裂型障碍和妄想性障碍",
            "F30-F39": "心境障碍（情感性障碍）",
            "F40-F48": "神经症性、应激相关及躯体形式障碍",
            "F50-F59": "与生理紊乱和躯体因素有关的行为综合征",
            "F60-F69": "成人人格和行为障碍",
            "F70-F79": "精神发育迟滞",
            "F80-F89": "心理发育障碍",
            "F90-F98": "儿童少年期行为和情绪障碍",
            "F99": "未特指的精神障碍"
        }
        
        for item in diagnoses_list:
            icd_code = item.get("icd_code", "")
            diagnosis = item.get("diagnosis", "")
            
            if not icd_code:
                continue
            
            # 提取大类代码（如F32, F20等）
            # 处理带小数点的代码（如F32.0）和不带小数点的代码（如F32）
            if '.' in icd_code:
                category_code = icd_code.split('.')[0]  # F32.0 -> F32
            else:
                category_code = icd_code  # F32 -> F32
            
            # 确定所属范围
            category_range = None
            category_name = "其他"
            
            # 提取数字部分用于范围判断
            try:
                code_num = int(category_code[1:])  # F32 -> 32
                
                if 0 <= code_num <= 9:
                    category_range = "F00-F09"
                elif 10 <= code_num <= 19:
                    category_range = "F10-F19"
                elif 20 <= code_num <= 29:
                    category_range = "F20-F29"
                elif 30 <= code_num <= 39:
                    category_range = "F30-F39"
                elif 40 <= code_num <= 48:
                    category_range = "F40-F48"
                elif 50 <= code_num <= 59:
                    category_range = "F50-F59"
                elif 60 <= code_num <= 69:
                    category_range = "F60-F69"
                elif 70 <= code_num <= 79:
                    category_range = "F70-F79"
                elif 80 <= code_num <= 89:
                    category_range = "F80-F89"
                elif 90 <= code_num <= 98:
                    category_range = "F90-F98"
                elif code_num == 99:
                    category_range = "F99"
                
                if category_range:
                    category_name = category_names.get(category_range, category_range)
            except (ValueError, IndexError):
                category_range = "其他"
                category_name = "其他"
            
            # 初始化大类
            if category_range not in grouped_data:
                grouped_data[category_range] = {
                    "range": category_range,
                    "name": category_name,
                    "subcategories": {}
                }
            
            # 初始化子类（如F32下的所有编码）
            if category_code not in grouped_data[category_range]["subcategories"]:
                grouped_data[category_range]["subcategories"][category_code] = {
                    "code": category_code,
                    "items": []
                }
            
            # 添加具体的诊断项
            grouped_data[category_range]["subcategories"][category_code]["items"].append({
                "icd_code": icd_code,
                "diagnosis": diagnosis
            })
        
        # 转换为列表格式并排序
        result = []
        for range_key in sorted(grouped_data.keys()):
            category = grouped_data[range_key]
            # 转换subcategories为列表
            subcategories_list = []
            for subcat_code in sorted(category["subcategories"].keys()):
                subcat = category["subcategories"][subcat_code]
                subcategories_list.append({
                    "code": subcat["code"],
                    "items": subcat["items"]
                })
            
            result.append({
                "range": category["range"],
                "name": category["name"],
                "subcategories": subcategories_list
            })
        
        return jsonify({
            "success": True,
            "data": result,
            "total_categories": len(result)
        })
        
    except Exception as e:
        print(f"获取分组诊断列表失败: {e}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


def main():
    """主函数"""
    import socket
    import os
    
    # 清除代理设置，避免影响本地服务
    for proxy_var in ["http_proxy", "https_proxy", "HTTP_PROXY", "HTTPS_PROXY", "all_proxy", "ALL_PROXY"]:
        if proxy_var in os.environ:
            del os.environ[proxy_var]
    
    print("=== Patient Agent 测试 UI 后端服务 ===")
    print(f"项目根目录: {project_root}")
    print(f"配置信息: {CONFIG}")
    
    # 加载患者数据
    if not load_patient_data():
        print("错误: 无法加载患者数据，程序退出")
        sys.exit(1)
    
    # 获取本机IP地址
    try:
        # 连接外部地址来获取本机IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        local_ip = "127.0.0.1"
    
    print(f"服务启动中...")
    print(f"前端静态文件目录: {app.static_folder}")
    print(f"")
    print(f"🌐 访问地址:")
    print(f"   本机访问: http://localhost:{CONFIG['backend_port']}")
    print(f"   局域网访问: http://{local_ip}:{CONFIG['backend_port']}")
    print(f"")
    print(f"📡 API接口:")
    print(f"   健康检查: http://{local_ip}:{CONFIG['backend_port']}/api/health")
    print(f"")
    print(f"💡 局域网内其他设备可通过 {local_ip}:{CONFIG['backend_port']} 访问此服务")
    
    # 启动Flask应用 - 配置为支持局域网访问
    app.run(
        host='0.0.0.0',  # 监听所有网络接口
        port=CONFIG["backend_port"],
        debug=False,     # 生产环境关闭debug
        threaded=True,
        use_reloader=False  # 避免重复启动
    )


if __name__ == "__main__":
    main()
